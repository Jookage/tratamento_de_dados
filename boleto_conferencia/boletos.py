import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ==========================================================
# ARQUIVO
# ==========================================================

ARQUIVO = "retorno_arquivo.TXT"

# ==========================================================
# EXPRESSÕES REGULARES
# ==========================================================

regex_data = re.compile(r"DATA ARQ - (\d{2}/\d{2}/\d{4})")

regex_registro = re.compile(
    r"""
    ^\s*999\s+
    (?P<agencia>\d{3})\s+
    (?P<conta>\d{16})\s+
    (?P<nome>.+?)
    \s+
    (?P<credito>\d[\d\.,]*)
    \s+
    (?P<debito>\d[\d\.,]*)
    \s+
    (?P<flag>\S+)
    \s+
    (?P<tipo>\S+)
    \s+
    (?P<sms>\S+)
    """,
    re.VERBOSE
)

# ==========================================================
# FUNÇÕES
# ==========================================================

def moeda(valor):

    valor = valor.replace(".", "").replace(",", ".")

    try:
        return float(valor)
    except:
        return 0.0

# ==========================================================
# LEITURA
# ==========================================================

dados = defaultdict(list)

data_atual = None

with open(ARQUIVO, encoding="latin1", errors="ignore") as arquivo:

    for linha in arquivo:

        # Procura DATA ARQ
        m_data = regex_data.search(linha)

        if m_data:
            data_atual = m_data.group(1)
            continue

        if data_atual is None:
            continue

        # Procura registros
        m = regex_registro.match(linha)

        if not m:
            continue

        registro = m.groupdict()

        # Remove DD
        if registro["tipo"] == "DD":
            continue

        dados[data_atual].append({

            "Agência": registro["agencia"],

            "Conta": registro["conta"],

            "Nome": registro["nome"].strip(),

            "Crédito": moeda(registro["credito"]),

            "Débito": moeda(registro["debito"]),

            "Flag": registro["flag"],

            "Tipo": registro["tipo"],

            "SMS": registro["sms"]

        })

# ==========================================================
# EXPORTAÇÃO
# ==========================================================

saida = Path(ARQUIVO).with_suffix(".xlsx")

with pd.ExcelWriter(saida, engine="openpyxl") as writer:

    for data, registros in dados.items():

        if not registros:
            continue

        df = pd.DataFrame(registros)

        nome_aba = data.replace("/", "-")

        df.to_excel(
            writer,
            sheet_name=nome_aba,
            index=False
        )

        ws = writer.sheets[nome_aba]

        # Cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Formato moeda
        for coluna in ("D", "E"):

            for cell in ws[coluna][1:]:

                cell.number_format = 'R$ #,##0.00'

        # Largura automática
        for i, coluna in enumerate(df.columns, start=1):

            largura = max(
                len(coluna),
                df[coluna].astype(str).map(len).max()
            ) + 3

            ws.column_dimensions[get_column_letter(i)].width = largura

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

# ==========================================================
# RESUMO
# ==========================================================

print("=" * 60)

total = 0

for data, registros in dados.items():

    print(f"{data}: {len(registros):,} registros")

    total += len(registros)

print("-" * 60)

print(f"TOTAL: {total:,} registros")

print(f"Arquivo salvo em:\n{saida.resolve()}")

print("=" * 60)